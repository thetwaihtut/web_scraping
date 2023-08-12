from bs4 import BeautifulSoup as bs
import requests
import openpyxl

excel=openpyxl.Workbook()
print(excel.sheetnames)
sheet=excel.active
sheet.title='Top 50 rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Rank','Movie Name','Year of Release','IMDB Rating'])

try:
    data=requests.get('https://www.imdb.com/search/title/?groups=top_250&sort=user_rating')
    data.raise_for_status()
    soup=bs(data.text,'lxml')
    movies=soup.find('div',class_='lister-list').find_all('div',class_='lister-item mode-advanced')
    for movie in movies:
        name=movie.find('h3',class_='lister-item-header').a.text
        rank=movie.find('h3',class_='lister-item-header').get_text(strip=True).split('.')[0]
        year1=movie.find('h3',class_='lister-item-header')
        year=year1.find('span',class_='lister-item-year text-muted unbold').text.strip('()')
        rating=movie.find('div',class_='ratings-bar').strong.text
        print(rank,name,year,rating)
        sheet.append([rank,name,year,rating])
        #break
except Exception as e:
    print(e)

excel.save('IMDB Movies Rating.xlsx')